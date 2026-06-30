import io
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

RATING_LABEL = {
    "fire": "🔥 Отлично",
    "ok": "✅ Хорошо",
    "warn": "⚠️ Риски",
    "bad": "❌ Нецелесообразно",
}

RATING_COLOR = {
    "fire": "C6EFCE",
    "ok": "CCFFCC",
    "warn": "FFEB9C",
    "bad": "FFC7CE",
}

COLUMNS = [
    ("Номер", 18),
    ("Название", 55),
    ("НМЦ, ₽", 18),
    ("Себестоимость, ₽", 20),
    ("Маржа, %", 12),
    ("Рейтинг", 20),
    ("Заказчик", 40),
    ("Регион", 22),
    ("Срок подачи", 16),
    ("Анализ Claude", 60),
    ("Ссылка", 55),
]


def generate_excel() -> bytes:
    """Генерирует Excel и возвращает байты файла."""
    from dotenv import load_dotenv
    load_dotenv()

    from app.database import get_tenders
    tenders = get_tenders(limit=500, only_analyzed=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Тендеры"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    for row_idx, t in enumerate(tenders, start=2):
        rating = t.get("rating") or ""
        row_color = RATING_COLOR.get(rating, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=row_color)

        values = [
            t.get("number", ""),
            t.get("title", ""),
            round(t.get("nmc") or 0),
            round(t.get("cost_estimate") or 0),
            round(t.get("margin_percent") or 0, 1),
            RATING_LABEL.get(rating, "—"),
            t.get("customer", ""),
            t.get("region", ""),
            t.get("end_date", ""),
            t.get("analysis_text", ""),
            t.get("url", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin
            cell.alignment = center if col_idx in (3, 4, 5, 6, 9) else left

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


if __name__ == "__main__":
    print("⏳ Генерирую Excel…", flush=True)
    xlsx = generate_excel()
    desktop = os.path.join(os.path.expanduser("~"), "Desktop", "Тендеры")
    os.makedirs(desktop, exist_ok=True)
    filename = f"tenders_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    path = os.path.join(desktop, filename)
    with open(path, "wb") as f:
        f.write(xlsx)
    print(f"✅ Сохранено: {path}", flush=True)
